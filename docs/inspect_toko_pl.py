import openpyxl

file = "docs/DATA_DARI_USER/DATA DARI USER UNTUK BUAT SISTEM/Lap Rugi laba 1 Jan sd 31 Maret 2026 (LAPORAN TRIWULAN TOKO).xlsx"
wb = openpyxl.load_workbook(file, read_only=True)
print("Toko P&L Sheets:", wb.sheetnames)

for name in wb.sheetnames:
    print(f"\n--- Sheet: {name} ---")
    sheet = wb[name]
    rows = list(sheet.iter_rows(values_only=True))
    for i, r in enumerate(rows[:20]):
        if any(x is not None for x in r):
            print(f"Row {i+1}: {[str(x)[:30] if x is not None else '' for x in r[:10]]}")
