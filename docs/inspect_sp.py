import openpyxl

file = "docs/DATA_DARI_USER/DATA DARI USER UNTUK BUAT SISTEM/LAPORAN BULANAN SIMPAN PINJAM.xlsx"
wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
print("Sheet Names:", wb.sheetnames)

for sheetname in wb.sheetnames:
    print(f"\n--- Sheet: {sheetname} ---")
    sheet = wb[sheetname]
    rows = list(sheet.iter_rows(values_only=True))
    non_empty_rows = [r for r in rows if any(x is not None for x in r)]
    print(f"Total non-empty rows: {len(non_empty_rows)}")
    for i, r in enumerate(non_empty_rows[:25]):
        display_row = [str(x)[:25] if x is not None else "" for x in r[:15]]
        print(f"Row {i+1}: {display_row}")
