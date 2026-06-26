import openpyxl
import os

files = [
    "docs/DATA_DARI_USER/DATA DARI USER UNTUK BUAT SISTEM/LAPORAN BULANAN SIMPAN PINJAM.xlsx",
    "docs/DATA_DARI_USER/DATA DARI USER UNTUK BUAT SISTEM/LAPORAN RAT 2025.xlsx"
]

for file in files:
    print(f"\n=========================================\nFILE: {file}")
    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        print(f"Sheets: {wb.sheetnames}")
        for sheetname in wb.sheetnames[:4]:
            print(f"\n--- Sheet: {sheetname} ---")
            sheet = wb[sheetname]
            rows = list(sheet.iter_rows(values_only=True))
            for i, row in enumerate(rows[:30]):
                if any(x is not None for x in row):
                    display_row = [str(x)[:30] if x is not None else "" for x in row[:12]]
                    print(f"Row {i+1}: {display_row}")
    except Exception as e:
        print(f"Error reading {file}: {e}")
