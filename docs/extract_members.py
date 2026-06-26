import openpyxl
import json

file = "docs/DATA_DARI_USER/DATA DARI USER UNTUK BUAT SISTEM/LAPORAN BULANAN SIMPAN PINJAM.xlsx"
wb = openpyxl.load_workbook(file, data_only=True)
sheet = wb['MEI 2026']

members = []
for r in range(4, sheet.max_row + 1):
    no = sheet.cell(r, 1).value
    name = sheet.cell(r, 2).value
    
    # Check if this row is a valid member row
    if no is not None and isinstance(no, int) and name is not None and str(name).strip() != "":
        name = str(name).strip()
        # Initial debt
        debt = sheet.cell(r, 3).value or 0
        # Installment, Interest, Penalty
        installment = sheet.cell(r, 4).value or 0
        interest = sheet.cell(r, 5).value or 0
        penalty = sheet.cell(r, 6).value or 0
        
        # Savings
        tab_pokok = sheet.cell(r, 7).value or 0
        tab_wajib = sheet.cell(r, 8).value or 0
        tab_sukarela = sheet.cell(r, 9).value or 0
        
        # Mandatory Savings Balances
        s_awal_wajib = sheet.cell(r, 11).value or 0
        wd_wajib = sheet.cell(r, 12).value or 0
        s_akhir_wajib = sheet.cell(r, 13).value or 0
        
        # Voluntary Savings Balances
        s_awal_sukarela = sheet.cell(r, 14).value or 0
        wd_sukarela = sheet.cell(r, 15).value or 0
        s_akhir_sukarela = sheet.cell(r, 16).value or 0
        
        # New Loan
        new_loan = sheet.cell(r, 17).value or 0
        provision = sheet.cell(r, 18).value or 0
        crk = sheet.cell(r, 19).value or 0
        
        members.append({
            "no": no,
            "name": name,
            "s_awal_hutang": debt,
            "angsuran": installment,
            "bunga": interest,
            "denda": penalty,
            "tab_pokok": tab_pokok,
            "tab_wajib": tab_wajib,
            "tab_sukarela": tab_sukarela,
            "s_awal_wajib": s_awal_wajib,
            "pengambilan_wajib": wd_wajib,
            "s_akhir_wajib": s_akhir_wajib,
            "s_awal_sukarela": s_awal_sukarela,
            "pengambilan_sukarela": wd_sukarela,
            "s_akhir_sukarela": s_akhir_sukarela,
            "pinjaman_baru": new_loan,
            "provisi": provision,
            "cad_resiko_kredit": crk
        })

print(f"Successfully extracted {len(members)} members.")
with open("docs/extracted_members.json", "w") as f:
    json.dump(members, f, indent=2)

# Summarize totals to verify
total_debt = sum(m["s_awal_hutang"] for m in members)
total_wajib = sum(m["s_akhir_wajib"] for m in members)
total_sukarela = sum(m["s_akhir_sukarela"] for m in members)
print(f"Total Initial Debt: {total_debt}")
print(f"Total Mandatory Savings (Ending): {total_wajib}")
print(f"Total Voluntary Savings (Ending): {total_sukarela}")
