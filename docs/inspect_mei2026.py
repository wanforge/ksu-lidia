import openpyxl
import json

file = "docs/DATA_DARI_USER/DATA DARI USER UNTUK BUAT SISTEM/LAPORAN BULANAN SIMPAN PINJAM.xlsx"
wb = openpyxl.load_workbook(file, data_only=True)
sheet = wb['MEI 2026']

print("Loaded sheet MEI 2026")
print("Max row:", sheet.max_row, "Max column:", sheet.max_column)

# Let's read first 10 rows to inspect headers
for r in range(1, 10):
    row_vals = [sheet.cell(r, c).value for c in range(1, 20)]
    print(f"Row {r}: {row_vals}")
