import openpyxl

file = "docs/DATA_DARI_USER/DATA DARI USER UNTUK BUAT SISTEM/LAPORAN BULANAN SIMPAN PINJAM.xlsx"
wb = openpyxl.load_workbook(file, read_only=True)
print("All Sheet Names:", wb.sheetnames)

for name in wb.sheetnames:
    print(f"Sheet: {name}, Class: {type(wb[name]).__name__}")
